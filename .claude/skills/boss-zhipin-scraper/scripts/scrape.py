#!/usr/bin/env python3
"""
BOSS直聘职位批量采集脚本
从搜索列表页采集职位详情，写入Excel文件。
"""

import subprocess
import json
import time
import re
import argparse
import sys
import os
from urllib.parse import quote

import openpyxl


# ── 13 extracted fields ──────────────────────────────────────────
HEADERS = [
    '岗位名称', '公司名称', '薪资', '经验', '学历',
    '岗位描述', 'HR', '工作地点', '福利',
    '公司基本信息', '公司介绍', '工商信息', '网址',
]

# ── Helpers ──────────────────────────────────────────────────────

def write_file(path, content):
    with open(path, 'w') as f:
        f.write(content)

def run_applescript(code):
    """Run AppleScript code and return stdout."""
    r = subprocess.run(['osascript', '-e', code], capture_output=True, text=True, timeout=30)
    return r.stdout.strip()

def run_js(js_code):
    """Execute JavaScript in Chrome active tab, return parsed JSON or None."""
    write_file('/tmp/_boss_scrape.js', js_code)
    as_code = '''
tell application "Google Chrome"
    tell active tab of front window
        set jsContent to read "/tmp/_boss_scrape.js"
        execute javascript jsContent
    end tell
end tell
'''
    write_file('/tmp/_boss_scrape.applescript', as_code)
    r = subprocess.run(['osascript', '/tmp/_boss_scrape.applescript'], capture_output=True, text=True, timeout=30)
    out = r.stdout.strip()
    if out and out != 'missing value':
        try:
            return json.loads(out)
        except json.JSONDecodeError:
            return None
    return None

def chrome_navigate(url):
    """Navigate Chrome active tab to URL."""
    code = f'tell application "Google Chrome"\n tell active tab of front window\n set URL to "{url}"\n end tell\nend tell'
    subprocess.run(['osascript', '-e', code], capture_output=True, timeout=15)

def chrome_get_url():
    """Get current Chrome URL."""
    code = 'tell application "Google Chrome"\n tell active tab of front window\n return URL\n end tell\nend tell'
    return run_applescript(code)


# ── Page interaction ─────────────────────────────────────────────

def scroll_to_load(target_count, max_scrolls=30):
    """Scroll the search results page to load enough job listings."""
    for _ in range(max_scrolls):
        urls = collect_job_urls()
        if len(urls) >= target_count:
            return urls
        run_js("window.scrollTo(0, document.body.scrollHeight); 'scrolled';")
        time.sleep(2)
    return collect_job_urls()

def collect_job_urls():
    """Collect all unique job detail URLs from search results page."""
    js = """
var links=document.querySelectorAll('a[href*="/job_detail/"]');
var seen={}; var r=[];
for(var i=0;i<links.length;i++){
    var h=links[i].href;
    if(h.indexOf('securityId')===-1 && h!=='https://www.zhipin.com/job_detail/'){
        var id=h.split('/').pop().split('.')[0];
        if(!seen[id] && id.length>5){
            seen[id]=true;
            r.push(h);
        }
    }
}
JSON.stringify(r);
"""
    data = run_js(js)
    return data if data else []


# ── Detail page extraction ───────────────────────────────────────

def extract_detail_page(url):
    """Navigate to detail page and extract all 13 fields."""
    chrome_navigate(url)
    time.sleep(3.5)

    js = """
var r={};
r.url=window.location.href;
var banner=document.querySelector('.job-banner');
if(banner){r.banner=banner.innerText;}
var tags=document.querySelector('.job-tags');
if(tags){r.tags=tags.innerText;}
var sider=document.querySelector('.job-sider');
if(sider){r.sider=sider.innerText;}
var detail=document.querySelector('.job-detail');
if(detail){r.detailFull=detail.innerText;}
JSON.stringify(r);
"""
    return run_js(js)

def parse_job(data):
    """Parse raw page data into structured job fields."""
    banner = data.get('banner', '')
    detail = data.get('detailFull', '')
    sider = data.get('sider', '')
    tags = data.get('tags', '')
    url = data.get('url', '')

    # ── Banner parsing: job name, salary, experience, education ──
    job_name = salary = exp = edu = ''
    lines = banner.strip().split('\n')
    for line in lines:
        line = line.strip()
        if line in ('招聘中', '最新', '感兴趣', '立即沟通', '继续沟通',
                     '完善在线简历', '新增附件简历', ''):
            continue
        m = re.search(r'(.+?)\s+(\d+[-–—]\d+[Kk]|\d+[-–—]\d+元/\w+|\d+[-–—]\d+元/时|\d+\.\d+[-–—]\d+[Kk])', line)
        if m:
            job_name = m.group(1).strip()
            salary = m.group(2).strip()
            break

    for line in lines:
        line = line.strip()
        if '台州' in line or '经验' in line:
            parts = line.split()
            for p in parts:
                if re.search(r'\d+-\d+年', p) and '日结' not in p:
                    exp = p
                elif p in ('经验不限', '在校/应届'):
                    exp = p
                elif p in ('大专', '本科', '硕士', '高中', '初中及以下', '学历不限'):
                    edu = p

    if not exp:
        for line in lines:
            if '经验不限' in line: exp = '经验不限'
            elif re.search(r'\d+-\d+年', line):
                m = re.search(r'(\d+-\d+年)', line)
                if m: exp = m.group(1)
    if not edu:
        for line in lines:
            if '学历不限' in line: edu = '学历不限'
            elif '大专' in line: edu = '大专'
            elif '本科' in line: edu = '本科'
            elif '高中' in line: edu = '高中'

    # ── Benefits from tags ──
    benefits = tags.strip() if tags else ''

    # ── Company basic info from sider ──
    sider_clean = sider.replace('公司基本信息\n\n', '').replace('公司基本信息\n', '').strip()
    sider_clean = sider_clean.replace('\n查看全部职位', '').replace('查看全部职位', '').strip()
    company = sider_clean.split('\n')[0].strip() if sider_clean else ''

    # ── Detail text parsing ──
    desc_text = detail[detail.find('职位描述')+4:].strip() if '职位描述' in detail else detail

    # HR
    hr_match = re.search(r'(\S+)\n刚刚活跃\n(.+?)\n·\n(.+?)\n', desc_text)
    hr = hr_match.group(1).strip() if hr_match else ''
    if (not company or company == job_name) and hr_match:
        company = hr_match.group(2).strip()

    # Job description (ends at HR or 竞争力分析/BOSS安全提示/公司介绍)
    desc_end = len(desc_text)
    for marker in ['\n竞争力分析', '\nBOSS 安全提示', '\n公司介绍']:
        pos = desc_text.find(marker)
        if 0 <= pos < desc_end:
            desc_end = pos
    if hr_match:
        pos = desc_text.find(hr_match.group(0))
        if 0 <= pos < desc_end:
            desc_end = pos
    job_desc = desc_text[:desc_end].strip()

    # Work address
    addr = ''
    am = re.search(r'工作地址\n(.+?)(?:\n|$)', desc_text)
    if am:
        addr = am.group(1).strip()

    # Company introduction
    intro = ''
    im = re.search(r'公司介绍\n((?:(?!\n查看全部\n).)*?)(?=\n查看全部\n|\n工商信息)', desc_text, re.DOTALL)
    if im:
        intro = im.group(1).strip()

    # Business registration info
    biz = ''
    bm = re.search(r'工商信息\n((?:(?!\n查看全部\n).)*?)(?=\n查看全部\n|\n工作地址)', desc_text, re.DOTALL)
    if bm:
        biz = bm.group(1).strip()

    return {
        '岗位名称': job_name, '公司名称': company, '薪资': salary,
        '经验': exp, '学历': edu, '岗位描述': job_desc, 'HR': hr,
        '工作地点': addr, '福利': benefits,
        '公司基本信息': sider_clean, '公司介绍': intro,
        '工商信息': biz, '网址': url,
    }


# ── Excel output ─────────────────────────────────────────────────

def write_excel(jobs, output_path):
    """Write or append job records to Excel file."""
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        existing_titles = set()
        for row in range(2, ws.max_row + 1):
            t = ws.cell(row, 1).value
            if t:
                existing_titles.add(t)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '职位详情'
        for ci, h in enumerate(HEADERS, 1):
            ws.cell(1, ci).value = h
        existing_titles = set()

    new_count = 0
    for job in jobs:
        if job['岗位名称'] in existing_titles:
            continue
        existing_titles.add(job['岗位名称'])
        row = ws.max_row + 1
        for ci, key in enumerate(HEADERS, 1):
            ws.cell(row, ci).value = job[key]
        new_count += 1

    wb.save(output_path)
    return new_count


# ── Main ─────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='BOSS直聘职位批量采集')
    parser.add_argument('--query', required=True, help='搜索关键词 (e.g. 数据标注师)')
    parser.add_argument('--city', required=True, help='城市编码 (e.g. 101210600 for 台州)')
    parser.add_argument('--count', type=int, default=20, help='采集数量 (default: 20)')
    parser.add_argument('--output', default='职位详情.xlsx', help='输出Excel路径 (default: 职位详情.xlsx)')
    args = parser.parse_args()

    query_encoded = quote(args.query)
    search_url = (
        f'https://www.zhipin.com/web/geek/jobs'
        f'?query={query_encoded}&city={args.city}&industry=&position='
    )

    print(f'>>> BOSS直聘职位采集 <<<')
    print(f'  关键词: {args.query}')
    print(f'  城市: {args.city}')
    print(f'  目标数量: {args.count}')
    print(f'  输出: {args.output}\n')

    # 1. Navigate to search results
    print('[1/4] 打开搜索页面...')
    chrome_navigate(search_url)
    time.sleep(4)

    # 2. Scroll to load enough jobs
    print('[2/4] 滚动加载职位列表...')
    urls = scroll_to_load(args.count)
    print(f'  收集到 {len(urls)} 个不重复职位URL')

    if len(urls) > args.count:
        urls = urls[:args.count]

    # 3. Visit each detail page and extract
    print(f'[3/4] 逐个提取详情 ({len(urls)} 个)...')
    all_jobs = []
    for i, url in enumerate(urls):
        name = url.split('/')[-1].split('.')[0][:20]
        print(f'  [{i+1}/{len(urls)}] {name}...', end=' ', flush=True)
        data = extract_detail_page(url)
        if data:
            job = parse_job(data)
            all_jobs.append(job)
            print(f'OK | {job["薪资"]} | {job["公司名称"]}')
        else:
            print('FAILED')
        time.sleep(0.3)

    # 4. Write to Excel
    print(f'[4/4] 写入Excel...')
    new_count = write_excel(all_jobs, args.output)
    print(f'  新增 {new_count} 条, 总计保存在 {args.output}')

    print(f'\n完成! 共采集 {new_count} 条职位信息.')


if __name__ == '__main__':
    main()
