#!/usr/bin/env python3
"""Import job listings from Excel into the database."""
import argparse
import sys

import openpyxl

from storage.db import init_db, get_session, insert_job
from storage.models import Job, Company, HRContact


def import_from_excel(excel_path: str):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Read headers from row 1
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    engine = init_db()
    session = get_session(engine)

    total = 0
    new = 0
    for row in range(2, ws.max_row + 1):
        data = {}
        for ci, key in enumerate(headers):
            val = ws.cell(row, ci + 1).value
            data[key] = str(val).strip() if val else ''

        if not data.get('网址'):
            continue

        total += 1
        job = insert_job(session, data)
        if job:
            new += 1
            if new % 10 == 0:
                session.commit()
                print(f'  已导入 {new} 条...')

    session.commit()

    job_count = session.query(Job).count()
    company_count = session.query(Company).count()
    hr_count = session.query(HRContact).count()

    print(f'\n结果:')
    print(f'  Excel 共 {total} 条, 新增 {new} 条, 跳过重复 {total - new} 条')
    print(f'  数据库: {job_count} 岗位, {company_count} 公司, {hr_count} HR')

    session.close()


def main():
    parser = argparse.ArgumentParser(description='导入 Excel 岗位数据到数据库')
    parser.add_argument('excel', help='Excel 文件路径')
    args = parser.parse_args()
    import_from_excel(args.excel)


if __name__ == '__main__':
    main()
